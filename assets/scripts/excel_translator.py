# -*- coding: utf-8 -*-
import os
import sys
import json
import re
import argparse
import shutil
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from deep_translator import GoogleTranslator

# Định nghĩa regex nhận diện chữ tiếng Trung
CHINESE_REGEX = re.compile(r'[\u4e00-\u9fff]')

def contains_chinese(text):
    if not isinstance(text, str):
        return False
    return bool(CHINESE_REGEX.search(text))

def translate_text_online(text, engine, target_lang="vi", api_key=None):
    if not text or not text.strip():
        return text
        
    try:
        if engine == "gemini" and api_key:
            import requests
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
            headers = {"Content-Type": "application/json"}
            prompt = f"Bạn là một biên dịch viên chuyên nghiệp. Hãy dịch đoạn văn bản sau đây sang tiếng Việt (hoặc ngôn ngữ mã \"{target_lang}\"). Chỉ trả về bản dịch, không giải thích gì thêm:\n\n{text}"
            data = {
                "contents": [{
                    "parts": [{"text": prompt}]
                }]
            }
            response = requests.post(url, headers=headers, json=data, timeout=10)
            if response.status_code == 200:
                result = response.json()
                translated = result['candidates'][0]['content']['parts'][0]['text'].strip()
                if translated.startswith('"') and translated.endswith('"'):
                    translated = translated[1:-1]
                return translated
            else:
                print(f"[WARN] Gemini API failed with status {response.status_code}, fallback to Google Free.", file=sys.stderr)
        
        return GoogleTranslator(source='auto', target=target_lang).translate(text)
    except Exception as e:
        print(f"[WARN] Translation error for '{text}': {str(e)}", file=sys.stderr)
        return text

def process_excel(src_path, dest_path, dict_data, patch_data, engine, target_lang, api_key):
    # Sao chép y hệt định dạng từ file gốc sang file mới
    shutil.copy2(src_path, dest_path)
    
    wb = load_workbook(dest_path)
    
    total_translated = 0
    total_dict_matches = 0
    total_patch_matches = 0
    total_online_translations = 0
    
    # 1. Chuẩn hóa từ điển
    normalized_dict = {}
    for k, v in dict_data.items():
        if k:
            normalized_dict[str(k).strip().lower()] = str(v).strip()
            
    # 2. Bước 1: Thu thập tất cả các chuỗi chữ tiếng Trung độc nhất cần dịch online
    unique_texts = set()
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_patches = patch_data.get(sheet_name, {})
        
        for r_idx in range(1, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell_coord = cell.coordinate
                val = cell.value
                
                if val is None or str(val).startswith('='):
                    continue
                
                val_str = str(val).strip()
                
                # Bỏ qua nếu có trong phần vá lỗi tọa độ
                if cell_coord in sheet_patches:
                    continue
                
                if contains_chinese(val_str):
                    val_lower = val_str.lower()
                    
                    # Thử xem có khớp hoàn toàn từ điển không
                    if val_lower in normalized_dict:
                        continue
                        
                    # Thử thế từ điển một phần
                    temp_str = val_str
                    for key_word, trans_word in normalized_dict.items():
                        if key_word in val_lower:
                            pattern = re.compile(re.escape(key_word), re.IGNORECASE)
                            temp_str = pattern.sub(trans_word, temp_str)
                            
                    # Nếu thế xong vẫn còn tiếng Trung thì cần dịch online
                    if contains_chinese(temp_str):
                        unique_texts.add(temp_str)
                    else:
                        # Chỉ cần từ điển là đủ
                        pass
                elif target_lang == 'vi' and not contains_chinese(val_str) and len(val_str) > 1:
                    # Hỗ trợ thêm dịch các cụm từ tiếng Anh chuyên ngành sang tiếng Việt nếu cần
                    if re.search(r'[a-zA-Z]', val_str):
                        unique_texts.add(val_str)
                        
    # 3. Bước 2: Dịch song song các chuỗi độc nhất bằng ThreadPoolExecutor (Tránh bị treo)
    translation_cache = {}
    unique_list = list(unique_texts)
    
    if unique_list:
        # Giới hạn 5 luồng song song để tránh bị Google block IP
        max_workers = min(5, len(unique_list))
        
        def translate_single(text):
            try:
                translated = translate_text_online(text, engine, target_lang, api_key)
                return text, translated
            except Exception:
                return text, text
                
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            results = executor.map(translate_single, unique_list)
            for text, translated in results:
                if translated and translated != text:
                    translation_cache[text] = translated

    # 4. Bước 3: Điền các bản dịch vào Workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_patches = patch_data.get(sheet_name, {})
        
        for r_idx in range(1, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell_coord = cell.coordinate
                val = cell.value
                
                if val is None or str(val).startswith('='):
                    continue
                
                val_str = str(val).strip()
                
                # LỚP 2: Vá lỗi tọa độ (Ưu tiên nhất)
                if cell_coord in sheet_patches:
                    cell.value = sheet_patches[cell_coord]
                    total_patch_matches += 1
                    total_translated += 1
                    continue
                
                if contains_chinese(val_str) or (target_lang == 'vi' and re.search(r'[a-zA-Z]', val_str) and len(val_str) > 1):
                    val_lower = val_str.lower()
                    
                    # LỚP 1: Khớp hoàn toàn từ điển
                    if val_lower in normalized_dict:
                        cell.value = normalized_dict[val_lower]
                        total_dict_matches += 1
                        total_translated += 1
                        continue
                        
                    # Dịch một phần bằng từ điển
                    temp_str = val_str
                    matched_dict = False
                    for key_word, trans_word in normalized_dict.items():
                        if key_word in val_lower:
                            pattern = re.compile(re.escape(key_word), re.IGNORECASE)
                            temp_str = pattern.sub(trans_word, temp_str)
                            matched_dict = True
                            
                    if matched_dict:
                        if not contains_chinese(temp_str):
                            cell.value = temp_str
                            total_dict_matches += 1
                            total_translated += 1
                            continue
                        else:
                            # Nếu còn chữ Trung, tiếp tục đối chiếu với cache dịch online
                            val_str = temp_str
                    
                    # LỚP 3: Lấy từ cache dịch online
                    if val_str in translation_cache:
                        cell.value = translation_cache[val_str]
                        total_online_translations += 1
                        total_translated += 1

    # LỚP 4: Quét kiểm tra chất lượng (Verification Scan)
    remaining_chinese_cells = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for r_idx in range(1, ws.max_row + 1):
            for c_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                val = cell.value
                if val and not str(val).startswith('=') and contains_chinese(str(val)):
                    remaining_chinese_cells.append({
                        "sheet": sheet_name,
                        "coordinate": cell.coordinate,
                        "value": str(val)
                    })
                    
    wb.save(dest_path)
    wb.close()
    
    return {
        "success": True,
        "total_translated": total_translated,
        "dict_matches": total_dict_matches,
        "patch_matches": total_patch_matches,
        "online_translations": total_online_translations,
        "remaining_chinese_count": len(remaining_chinese_cells),
        "remaining_cells": remaining_chinese_cells[:20]
    }

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="JA Excel Translator Tool")
    parser.add_argument("--src", required=True, help="Path to source Excel file")
    parser.add_argument("--dest", required=True, help="Path to destination Excel file")
    parser.add_argument("--dict", default="{}", help="JSON string of translation dictionary")
    parser.add_argument("--patches", default="{}", help="JSON string of coordinate patching")
    parser.add_argument("--engine", default="google", help="Translation engine (google, gemini)")
    parser.add_argument("--lang", default="vi", help="Target language code")
    parser.add_argument("--key", default="", help="API Key for translation engine")
    
    args = parser.parse_args()
    
    try:
        try:
            dict_data = json.loads(args.dict)
        except Exception:
            dict_data = {}
            
        try:
            patch_data = json.loads(args.patches)
        except Exception:
            patch_data = {}
            
        result = process_excel(
            src_path=args.src,
            dest_path=args.dest,
            dict_data=dict_data,
            patch_data=patch_data,
            engine=args.engine,
            target_lang=args.lang,
            api_key=args.key
        )
        print(json.dumps(result, ensure_ascii=False))
    except Exception as e:
        error_res = {
            "success": False,
            "error": str(e)
        }
        print(json.dumps(error_res, ensure_ascii=False))
        sys.exit(1)
